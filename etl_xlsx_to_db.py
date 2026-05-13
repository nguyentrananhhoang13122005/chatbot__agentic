"""
ETL Pipeline: Đọc toàn bộ file XLSX điểm chuẩn → Xuất ra CSV chuẩn hóa.
Quét thư mục data/3. Điểm chuẩn năm 2025/, đọc tất cả sheet trong mỗi file,
tự động xác định header, chuẩn hóa schema và gộp thành 1 database duy nhất.
"""
import sys
import os
import re
import glob
import pandas as pd
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

# === CONFIG ===
XLSX_DIR = r"data/3. Điểm chuẩn năm 2025-20260507T075825Z-3-001/3. Điểm chuẩn năm 2025"
OUTPUT_CSV = "data/data_diem_chuan_verified.csv"

# === SCHEMA CHUẨN ===
STANDARD_COLS = ["Trường", "Mã ngành", "Tên ngành", "Năm", "Phương thức xét tuyển", "Điểm chuẩn", "Chỉ tiêu", "Tổ hợp môn", "Ghi chú"]

# === BẢNG ÁNH XẠ MÃ TRƯỜNG → TÊN ĐẦY ĐỦ (cho 119 file chỉ có mã) ===
SCHOOL_CODE_MAP = {
    "GDU": "Đại học Công nghệ Thông tin Gia Định", "GHA": "Đại học Gia Định", "GNT": "Đại học Công nghệ Giao thông Vận tải",
    "GSA": "Đại học Giao thông Vận tải TPHCM", "GTA": "Đại học Giao thông Vận tải", "GTS": "Đại học Giao thông Vận tải cơ sở 2",
    "HBT": "Học viện Báo chí và Tuyên truyền", "HCA": "Học viện Cảnh sát Nhân dân", "HCB": "Học viện Chính trị CAND",
    "HCH": "Đại học Công đoàn", "HCN": "Đại học Công nghiệp Hà Nội", "HCP": "Học viện Chính sách và Phát triển",
    "HCS": "Học viện Cơ sở", "HDT": "Đại học Điện lực", "HEH": "Đại học Kinh tế Quốc dân",
    "HFH": "Đại học Tài chính Ngân hàng Hà Nội", "HGH": "Đại học Hà Nội", "HHA": "Đại học Bách khoa Hà Nội",
    "HHK": "Đại học Hàng hải Việt Nam", "HHT": "Đại học Hà Tĩnh", "HIU": "Đại học Quốc tế Hồng Bàng",
    "HLU": "Đại học Luật Hà Nội", "HNM": "Đại học Nội vụ Hà Nội", "HPN": "Đại học Hải Phòng",
    "HPU": "Đại học Quản lý và Công nghệ Hải Phòng", "HQH": "Đại học Mở Hà Nội", "HQT": "Đại học Mở TPHCM",
    "HSU": "Đại học Hoa Sen", "HTA": "Học viện Kỹ thuật Mật mã", "HTC": "Đại học Hồng Đức",
    "HTN": "Đại học Hùng Vương", "HVA": "Học viện An ninh Nhân dân", "HVC": "Học viện Chính trị Quốc gia HCM",
    "HVD": "Học viện Dân tộc", "HVN": "Học viện Nông nghiệp Việt Nam", "HVQ": "Học viện Quân y",
    "HYD": "Đại học Y Hà Nội", "IUH": "Đại học Công nghiệp TPHCM", "KCC": "Đại học Kiến trúc TPHCM",
    "KCN": "Đại học Kiến trúc Hà Nội", "KGH": "Đại học Kinh tế - ĐH Huế", "KHA": "Đại học Kinh tế Quốc dân",
    "KMA": "Học viện Kỹ thuật Mật mã", "KQH": "Đại học Kinh tế Quốc dân", "KSA": "Đại học Kinh tế TPHCM",
    "KSV": "Đại học Khoa học Tự nhiên - ĐHQG TPHCM", "KTA": "Đại học Kiến trúc Đà Nẵng",
    "KTD": "Đại học Kinh tế Đà Nẵng", "KTS": "Đại học Kiến trúc TPHCM",
    "LAH": "Đại học Lâm nghiệp", "LBH": "Đại học Lao động Xã hội", "LBS": "Đại học Lao động Xã hội CS2",
    "LCH": "Đại học Luật TPHCM", "LDA": "Học viện Lục quân", "LNA": "Đại học Lạc Hồng",
    "LNH": "Đại học Lâm nghiệp Hà Nội", "LNS": "Đại học Lâm nghiệp CS2", "LPH": "Đại học Phòng cháy Chữa cháy",
    "LPS": "Đại học Phòng cháy Chữa cháy CS2",
    "MBS": "Đại học Mỏ - Địa chất", "MDA": "Đại học Mỹ thuật Công nghiệp", "MHN": "Đại học Mỹ thuật Việt Nam",
    "MIT": "Đại học Công nghệ Miền Đông", "MTC": "Đại học Mỹ thuật TPHCM", "MTH": "Đại học Mỏ - Địa chất",
    "MTS": "Đại học Mỹ thuật TPHCM", "MTU": "Đại học Mỏ - Địa chất",
    "NHB": "Đại học Ngân hàng TPHCM", "NHF": "Đại học Ngoại thương", "NHH": "Đại học Ngoại thương Hà Nội",
    "NHP": "Đại học Ngoại thương CS2", "NHS": "Đại học Ngoại thương TPHCM",
    "NLG": "Đại học Nông Lâm TPHCM", "NLN": "Đại học Nông Lâm Bắc Giang", "NLS": "Đại học Nông Lâm - ĐH Huế",
    "NQH": "Đại học Nha Trang", "NTH": "Đại học Nguyễn Tất Thành", "NTS": "Đại học Nguyễn Tất Thành CS2",
    "NTT": "Đại học Nguyễn Tất Thành", "NTU": "Đại học Nha Trang", "NVH": "Đại học Nội vụ Hà Nội", "NVS": "Đại học Nội vụ CS2",
    "PBH": "Đại học Phenikaa", "PCH": "Học viện Phòng không - Không quân", "PCH1": "Học viện Phòng không - Không quân DS",
    "PCS": "Học viện Phòng không CS2", "PCS1": "Học viện Phòng không CS2 DS",
    "PKA": "Đại học Phenikaa", "PKH": "Đại học Phú Xuân",
    "PVU": "Đại học Công nghệ GTVT - Phân hiệu Vĩnh Phúc",
    "QHD": "Đại học Khoa học Tự nhiên - ĐHQG Hà Nội", "QHE": "Đại học Kinh tế - ĐHQG Hà Nội",
    "QHF": "Đại học Ngoại ngữ - ĐHQG Hà Nội", "QHI": "Đại học Công nghệ - ĐHQG Hà Nội",
    "QHK": "Khoa Y Dược - ĐHQG Hà Nội", "QHL": "Đại học Luật - ĐHQG Hà Nội",
    "QHQ": "Đại học Quốc tế - ĐHQG TPHCM", "QHS": "Đại học KHXH&NV - ĐHQG Hà Nội",
    "QHT": "Đại học Giáo dục - ĐHQG Hà Nội", "QHX": "Đại học KHXH&NV - ĐHQG Hà Nội",
    "QHY": "Khoa Quốc tế - ĐHQG Hà Nội",
    "QSB": "Đại học Bách khoa - ĐHQG TPHCM", "QSC": "Đại học Công nghệ Thông tin - ĐHQG TPHCM",
    "QSK": "Đại học Khoa học Tự nhiên - ĐHQG TPHCM", "QSP": "Đại học KHXH&NV - ĐHQG TPHCM",
    "QSQ": "Đại học Quốc tế - ĐHQG TPHCM", "QST": "Đại học Kinh tế Luật - ĐHQG TPHCM",
    "QSX": "Đại học An Giang - ĐHQG TPHCM", "QSY": "Khoa Y - ĐHQG TPHCM",
    "SDU": "Đại học Sân khấu Điện ảnh Hà Nội", "SGD": "Đại học Sài Gòn",
    "SIU": "Đại học Quốc tế Sài Gòn", "SKD": "Đại học Sư phạm Kỹ thuật Đà Nẵng",
    "SKH": "Đại học Sư phạm Kỹ thuật Hưng Yên", "SKN": "Đại học Sư phạm Kỹ thuật Nam Định",
    "SKV": "Đại học Sư phạm Kỹ thuật Vinh", "SNH": "Đại học Sư phạm Nghệ thuật Trung ương",
    "SP2": "Đại học Sư phạm Hà Nội 2", "SPD": "Đại học Sư phạm Đà Nẵng",
}


def extract_school_name(filename: str) -> str:
    """Trích xuất tên trường từ tên file XLSX, sử dụng bảng ánh xạ cho file code-only."""
    base = os.path.splitext(os.path.basename(filename))[0]
    
    # Pattern 1: Full name "124. FPT_ĐH FPT 2025" hoặc "249.SPH-Đại Học Sư Phạm Hà Nội"
    m = re.match(r'^\d+[\.\s]+[A-Z0-9]{2,5}[\._\-\s]+(.+?)(?:\s*2025)?$', base)
    if m:
        name = m.group(1).strip()
        name = name.replace('_', ' ').strip()
        name = re.sub(r'\s+', ' ', name)
        if len(name) > 3:  # Tên thật sự, không phải mã
            return name
    
    # Pattern 2: Code-only "125.GDU" → tra bảng ánh xạ
    m = re.match(r'^\d+\.([A-Z0-9]{2,5})$', base)
    if m:
        code = m.group(1)
        return SCHOOL_CODE_MAP.get(code, code)
    
    # Fallback: bỏ số đầu, giữ phần còn lại
    base = re.sub(r'^\d+[\.\s]*', '', base).strip()
    base = re.sub(r'\s*2025\s*$', '', base).strip()
    return base if base else os.path.basename(filename)


def detect_header_row(ws) -> int:
    """Tìm dòng header chứa 'Mã ngành' hoặc 'Tên ngành'."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_text = " ".join([str(c).lower() for c in row if c])
        if 'mã ngành' in row_text or 'tên ngành' in row_text:
            return i
    return -1


def detect_method_from_sheet(sheet_title: str, first_row_text: str) -> str:
    """Xác định phương thức xét tuyển từ tên sheet hoặc dòng đầu tiên."""
    combined = (sheet_title + " " + first_row_text).lower()

    if 'đgnl' in combined or 'đánh giá năng lực' in combined:
        if 'hn' in combined or 'hà nội' in combined or 'hcm' not in combined:
            return "Xét điểm ĐGNL ĐHQG Hà Nội"
        return "Xét điểm ĐGNL ĐHQG TPHCM"
    if 'tư duy' in combined or 'đgtd' in combined:
        return "Xét điểm Đánh giá Tư duy"
    if 'học bạ' in combined or 'hb' in combined:
        return "Xét điểm Học bạ THPT"
    if 'thpt' in combined or 'thi tốt' in combined or 'tốt nghiệp' in combined:
        return "Xét điểm thi THPT"
    if 'chứng chỉ' in combined or 'quốc tế' in combined or 'ielts' in combined or 'sat' in combined:
        return "Xét chứng chỉ quốc tế"
    if 'xét tuyển sớm' in combined or 'sớm' in combined:
        return "Xét tuyển sớm"
    if 'kết hợp' in combined:
        return "Xét tuyển kết hợp"
    # Default
    return sheet_title.strip() if sheet_title.strip() else "Không xác định"


def map_columns(header_row: list) -> dict:
    """Map tên cột thực tế về tên cột chuẩn."""
    col_map = {}
    for i, col in enumerate(header_row):
        if col is None:
            continue
        col_lower = str(col).lower().strip()
        if 'mã ngành' in col_lower or 'mã' == col_lower:
            col_map['Mã ngành'] = i
        elif 'tên ngành' in col_lower or ('ngành' in col_lower and 'mã' not in col_lower):
            col_map['Tên ngành'] = i
        elif 'điểm chuẩn' in col_lower or 'điểm trúng' in col_lower or ('điểm' in col_lower and 'ghi' not in col_lower):
            col_map['Điểm chuẩn'] = i
        elif 'tổ hợp' in col_lower or 'khối' in col_lower:
            col_map['Tổ hợp môn'] = i
        elif 'chỉ tiêu' in col_lower:
            col_map['Chỉ tiêu'] = i
        elif 'ghi chú' in col_lower:
            col_map['Ghi chú'] = i
    return col_map


def process_one_file(filepath: str) -> list:
    """Xử lý 1 file XLSX, trả về list of dicts."""
    school_name = extract_school_name(filepath)
    records = []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as e:
        print(f"  ⚠️ Không mở được: {os.path.basename(filepath)} → {e}")
        return []

    for ws in wb.worksheets:
        # Lấy text dòng 1 để xác định phương thức
        first_row_vals = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            first_row_vals = [str(c) if c else "" for c in row]
        first_row_text = " ".join(first_row_vals)

        method = detect_method_from_sheet(ws.title, first_row_text)

        # Tìm header row
        header_idx = -1
        header_vals = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            row_text = " ".join([str(c).lower() for c in row if c])
            if 'mã ngành' in row_text or ('tên ngành' in row_text and 'stt' in row_text):
                header_idx = i
                header_vals = list(row)
                break

        if header_idx < 0:
            continue

        col_map = map_columns(header_vals)
        if 'Mã ngành' not in col_map and 'Tên ngành' not in col_map:
            continue

        # Đọc data rows
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            row_vals = list(row)

            # Skip hàng trống
            ma_idx = col_map.get('Mã ngành')
            ten_idx = col_map.get('Tên ngành')
            diem_idx = col_map.get('Điểm chuẩn')

            ma = str(row_vals[ma_idx]).strip() if ma_idx is not None and ma_idx < len(row_vals) and row_vals[ma_idx] else ""
            ten = str(row_vals[ten_idx]).strip() if ten_idx is not None and ten_idx < len(row_vals) and row_vals[ten_idx] else ""
            diem = row_vals[diem_idx] if diem_idx is not None and diem_idx < len(row_vals) else None

            # Skip dòng header lặp hoặc dòng rỗng
            if not ten or ten.lower() in ('none', 'nan', 'tên ngành', ''):
                continue
            if ma.lower() in ('none', 'nan', 'stt', 'mã ngành', 'mã'):
                continue

            # Parse điểm chuẩn
            diem_val = None
            if diem is not None:
                try:
                    diem_str = str(diem).replace(',', '.').strip()
                    diem_val = float(diem_str)
                except ValueError:
                    diem_val = str(diem).strip()

            to_hop_idx = col_map.get('Tổ hợp môn')
            chi_tieu_idx = col_map.get('Chỉ tiêu')
            ghi_chu_idx = col_map.get('Ghi chú')

            records.append({
                "Trường": school_name,
                "Mã ngành": ma,
                "Tên ngành": ten,
                "Năm": 2025,
                "Phương thức xét tuyển": method,
                "Điểm chuẩn": diem_val,
                "Chỉ tiêu": row_vals[chi_tieu_idx] if chi_tieu_idx and chi_tieu_idx < len(row_vals) else None,
                "Tổ hợp môn": row_vals[to_hop_idx] if to_hop_idx and to_hop_idx < len(row_vals) else None,
                "Ghi chú": row_vals[ghi_chu_idx] if ghi_chu_idx and ghi_chu_idx < len(row_vals) else None,
            })

    wb.close()
    return records


def main():
    print("=" * 60)
    print("🚀 ETL Pipeline: XLSX → Verified CSV Database")
    print("=" * 60)

    # Tìm tất cả file XLSX
    xlsx_files = glob.glob(os.path.join(XLSX_DIR, "*.xlsx"))
    print(f"\n📁 Tìm thấy {len(xlsx_files)} file XLSX trong thư mục.")

    all_records = []
    success = 0
    failed = 0

    for i, fpath in enumerate(sorted(xlsx_files), 1):
        fname = os.path.basename(fpath)
        records = process_one_file(fpath)
        if records:
            all_records.extend(records)
            success += 1
            print(f"  ✅ [{i}/{len(xlsx_files)}] {fname[:50]}... → {len(records)} ngành")
        else:
            failed += 1
            print(f"  ❌ [{i}/{len(xlsx_files)}] {fname[:50]}... → Không trích xuất được")

    # Tạo DataFrame và xuất CSV
    df = pd.DataFrame(all_records, columns=STANDARD_COLS)

    # Clean up
    df = df.dropna(subset=["Tên ngành"])
    df["Điểm chuẩn"] = pd.to_numeric(df["Điểm chuẩn"], errors="coerce")
    df = df.drop_duplicates(subset=["Trường", "Mã ngành", "Tên ngành", "Phương thức xét tuyển"], keep="first")

    # Xuất CSV
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    print(f"\n{'=' * 60}")
    print(f"📊 KẾT QUẢ ETL:")
    print(f"  ✅ Thành công: {success} file")
    print(f"  ❌ Thất bại:   {failed} file")
    print(f"  📝 Tổng ngành: {len(df)} dòng")
    print(f"  🏫 Số trường:  {df['Trường'].nunique()} trường duy nhất")
    print(f"  💾 Đã lưu:     {OUTPUT_CSV}")
    print(f"{'=' * 60}")

    # Hiển thị sample
    print(f"\n🔍 Sample 10 dòng đầu:")
    print(df.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
